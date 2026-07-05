# -*- coding: utf-8 -*-
"""Testy pro POST /api/add-sarze — přidání položky do katalogu šarží.

Seam: app.py drží SARZE_PATH a matcher jako modulové globály a add_sarze
z nich čte až za běhu. Testy je přes monkeypatch přesměrují na tmp kopii
minimálního sešitu — reálná data/sarze.xlsx zůstává nedotčená a monkeypatch
po každém testu vrátí původní globály.
"""
import openpyxl
import pytest

import app as app_module
from plant_matcher import PlantMatcher

HEADER = ('Název', 'Kód', 'Země')
SEED_ROW = ('Acer palmatum', '25-T1', 'CZ')


@pytest.fixture
def catalog(tmp_path, monkeypatch):
    path = tmp_path / 'sarze.xlsx'
    wb = openpyxl.Workbook()
    wb.active.append(list(HEADER))
    wb.active.append(list(SEED_ROW))
    wb.save(path)

    monkeypatch.setattr(app_module, 'SARZE_PATH', path)
    monkeypatch.setattr(app_module, 'matcher', PlantMatcher(str(path)))
    monkeypatch.setitem(app_module.app.config, 'TESTING', True)
    return path


@pytest.fixture
def client(catalog):
    with app_module.app.test_client() as c:
        yield c


def _rows(path):
    ws = openpyxl.load_workbook(path).active
    return list(ws.iter_rows(values_only=True))


def test_add_appends_one_trimmed_row_and_returns_exact_match(client, catalog):
    resp = client.post('/api/add-sarze', json={
        'name': '  Quercus robur  ', 'code': ' 25-X9 '})

    assert resp.status_code == 200
    data = resp.get_json()
    assert data['match_type'] == 'exact'
    assert data['code'] == '25-X9'
    assert data['country'] == 'CZ'                    # země vynechána → CZ
    assert data['passport_name'] == 'Quercus robur'
    # Přesně JEDEN nový řádek, hodnoty ořezané: [name, code, country]
    assert _rows(catalog) == [HEADER, SEED_ROW,
                              ('Quercus robur', '25-X9', 'CZ')]


def test_added_name_found_exact_by_search_plant(client):
    """Úspěch je viditelný přes matcher: následné /api/search-plant
    najde přidaný název jako přesnou shodu (globální matcher byl obnoven)."""
    client.post('/api/add-sarze',
                json={'name': 'Quercus robur', 'code': '25-X9'})

    resp = client.post('/api/search-plant', json={'query': 'Quercus robur'})

    assert resp.status_code == 200
    data = resp.get_json()
    assert data['match_type'] == 'exact'
    assert data['code'] == '25-X9'


def test_country_is_trimmed_and_stored(client, catalog):
    resp = client.post('/api/add-sarze', json={
        'name': 'Pinus nigra', 'code': '25-P1', 'country': '  DE  '})

    assert resp.status_code == 200
    assert resp.get_json()['country'] == 'DE'
    assert _rows(catalog)[-1] == ('Pinus nigra', '25-P1', 'DE')


@pytest.mark.parametrize('payload', [
    {},                                        # obojí chybí
    {'name': 'Quercus robur'},                 # chybí kód
    {'code': '25-X9'},                         # chybí název
    {'name': '   ', 'code': '25-X9'},          # název jen mezery
    {'name': 'Quercus robur', 'code': '   '},  # kód jen mezery
], ids=['empty', 'missing-code', 'missing-name', 'blank-name', 'blank-code'])
def test_invalid_input_400_leaves_catalog_untouched(client, catalog, payload):
    before = catalog.read_bytes()

    resp = client.post('/api/add-sarze', json=payload)

    assert resp.status_code == 400
    assert 'error' in resp.get_json()
    assert catalog.read_bytes() == before      # soubor bajt po bajtu netknutý
