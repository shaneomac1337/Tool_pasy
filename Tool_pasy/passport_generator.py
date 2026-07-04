import io
import base64
import tempfile
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path
from datetime import date


# EU vlajka zakódovaná jako base64 – nevyžaduje žádný externí soubor
_EU_FLAG_B64 = (
    'iVBORw0KGgoAAAANSUhEUgAAASwAAADICAIAAADdvUsCAAAFxElEQVR4nO3dUVLjOBSG0TA1O5pF'
    'sSZWyzxABcoxtmLL/nOtc16HdHuEPl/ATfR2++/jBuT8k74AGJ0IIUyEECZCCBMhhIkQwkQIYSKE'
    'MBFCmAghTIQQJkIIEyGEiRDCRAhhIoQwEUKYCCFMhBAmQggTIYSJEMJECGEihDARlvf58Z6+BHYR'
    'IYSJsLavMWgYliZCCBNhYQbgNYgQwkRY1WQMmop1vTmf8MXtr+vt3af4pZmEh9tZ0c6E9hdoxh5N'
    'hAVsDskMLOHf9AVcXK8x8pVT+5/WKz9j8AQm4Rn6ptjrw9pJ8VAiLGY1MF+CliPCA1UfINWvvwoR'
    'FrMahnLKEeFRqj9Mr379hfjp6F7tu/Ovj9zzXdz9tZsjyV4/N/9ippdtDWzYvve/6PG1C/+p/Y99'
    'ivy6EGE3z+7jvgVOPuaEDhXYiwg7a9nKm7fv58d7y2sbP+yv165+jPz6EmF/y/v49Xdw9esvx09H'
    'IUyEECbCzqo/TK9+/RWJ8HDVv4mqfv2vT4QHenv/+NrBRfdx9euvQoRHmWzc+4auovr1F+IRRU8t'
    'D8o3P0w/QfXrL8ok7Kz67/tVv/6KTEIIMwkhTIR/8kCMc4iQM7ijLRDhPEeOcRoRcjh3tGUihDAR'
    'zvh9z3b/3slirhIhhIlw6vFu7f69mcVsMfS/mHH0Xy9Wco+hJ2H86L/LsJJ7DD0Jv2y4iw++aRZY'
    'zA1E+O38o/+uyko+S4Q/VnePTdPOYrYb+nvCieVtYdM8xWK2EyGEOfDH8ldQHnA9xWK2EyGEOZ9w'
    '3uy5f3sOWhnKZNDtP0Tx2kzCGb9L81Z/e0xWz0rO8oji2+o7+e05gnMohx6ieEkm4Y/Vn6rbNI1W'
    '18pK/mYSQphJCGEihDARQpgIIUyEECZCCBMhhIkQwkQIYSKEMBH6/RrCREiYm+DoEdoBxI0e4Rcp'
    'pji68CZCiBs6wsFvwHHW/8vQEcIrGDfCyW3YXflk1v/u+m9v4ei8LOu/6vqT0NF5WdZ/1fUjvO34'
    'RI6wA05g/Zdd/8vR3xydl2X9Zw0xCe8aP7VD7YAzWf9ZY0V4a/gEj7YDTmb9Hw0XIbya4SJc/bZk'
    '5AdWJ7D+j4aLEF7N6OcTOjovy/rfRp6Ejs7Lsv53gz4ndHRehPWfNdwkdHRelvV/NNYkhBc03CSE'
    'VyNCCBMhhIkQwkQIYSKEMBFCmAghTIQQJkIIE+GfRv7lGs4kQs7gjrZAhPMc2cVpRMjh3NGWiRDC'
    'RDjj9z3b/Xsni7lKhBAmwqnHu7X792YWs8XQb2/h6LxerOQeQ09CR+f1YiX3GHoSftlwFx980yyw'
    'mBuI8Juj83qxks8S4Y/V3WPTtLOY7Yb+nnDCm9J2ZDHbiRDCRPhj+SsoD7ieYjHbiRDCRj+f8C+z'
    '5+Z9frz7ZqbFZNA5hHCZSThjcm6e8DZzCGELjyi+rR6L13K2HjeHED7PJPyx+lN1m6aRQwifYhJC'
    'mEkIYSKEMBEervTP5UtffBUihDARHqv0u/2VvvhCRAhhIjzQZWbIZf5HXpMIIUyER5lMj3LDpPr1'
    'FyJCCPPP1vaq/pab1a//AkTYx7at3GX7dvktx+D148vRPjZsx5fawdWvvzSTsLOWkdJx+3b/LceT'
    'r5+bSdjd6gZ98R1c/forEmFhHhtcgwghTISdrU6nXuProIfpp10/d97y8HBv7x97Nm77a//6yD3fxe'
    '28eFqYhAe6v9/Rzgx2XsPO13qHq6OJ8CiTjbtnK0de+Hj92/4oVnlO2FPL22nuecvNo4/+O/r6mW'
    'USdnboc7bG1x76V8ivO5OwnuV5KJJyTEIIE2ExnuNdjwghzMP62hz9dwEmYVWO/rsMPx2txNF/l2QS'
    'FuPov+sxCSHMJIQwEUKYCCFMhBAmQggTIYSJEMJECGEihDARQpgIIUyEECZCCBMhhIkQwkQIYSKE'
    'MBFCmAghTIQQJkIIEyGEiRDCRAhhIoQwEULY/8kFYm/ZtGj3AAAAAElFTkSuQmCC'
)


def _get_flag_image_path() -> str:
    """Vrátí cestu k souboru s EU vlajkou.
    Preferuje eu_flag.png vedle tohoto souboru, jinak použije base64 zálohu."""
    local_flag = Path(__file__).parent / 'eu_flag.png'
    if local_flag.exists():
        return str(local_flag)
    # záloha: dekóduj z base64 do dočasného souboru
    data = base64.b64decode(_EU_FLAG_B64)
    tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
    tmp.write(data)
    tmp.flush()
    tmp.close()
    return tmp.name


def generate_excel(invoices: list, output_dir: Path) -> Path:
    """
    Vygeneruje Excel soubor s pasy — přesný formát dle vzoru.

    invoices: seznam dict {
        'number':  str,
        'plants':  list[dict] { 'passport_name', 'code', 'country' }
    }
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"pasy_{date.today().isoformat()}.xlsx"
    output_path = output_dir / filename

    # Prepare flag image once for all sheets
    flag_path = _get_flag_image_path()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # odeber výchozí prázdný list

    for inv in invoices:
        plants = inv.get('plants', [])
        if not plants:
            continue
        sheet_name = str(inv['number'])[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_passport_sheet(ws, plants, flag_path)

    if not wb.sheetnames:
        ws = wb.create_sheet("Prázdné")

    wb.save(str(output_path))

    # Clean up temp flag file (pouze pokud jsme vytvořili dočasný soubor, ne eu_flag.png)
    local_flag = str(Path(__file__).parent / 'eu_flag.png')
    if flag_path != local_flag:
        try:
            import os
            os.unlink(flag_path)
        except Exception:
            pass

    return output_path


def _add_side(cell, left=None, right=None, top=None, bottom=None):
    """Přidá strany rámečku, zachová existující."""
    b = cell.border
    cell.border = Border(
        left   = left   if left   is not None else b.left,
        right  = right  if right  is not None else b.right,
        top    = top    if top    is not None else b.top,
        bottom = bottom if bottom is not None else b.bottom,
    )


def _outline(ws, min_col, min_row, max_col, max_row, side):
    """Orámuje obdélník (i přes sloučené buňky) bez vnitřních čar."""
    for col in range(min_col, max_col + 1):
        _add_side(ws.cell(min_row, col), top=side)
        _add_side(ws.cell(max_row, col), bottom=side)
    for row in range(min_row, max_row + 1):
        _add_side(ws.cell(row, min_col), left=side)
        _add_side(ws.cell(row, max_col), right=side)


def _write_passport_sheet(ws, plants: list, flag_path: str):
    """Zapíše jeden list přesně dle vzoru."""

    # ── Šířky sloupců (dle vzoru) ─────────────────────────────
    ws.column_dimensions['A'].width = 13.0
    ws.column_dimensions['B'].width = 4.71
    ws.column_dimensions['C'].width = 48.71
    ws.column_dimensions['D'].width = 21.14
    ws.column_dimensions['E'].width = 23.86
    ws.column_dimensions['F'].width = 4.29

    # ── Výšky řádků — jednotná 15.75 ─────────────────────────
    # Header (12) + data rows + 1 empty closing row
    total_rows = 13 + len(plants) + 1
    for r in range(1, total_rows + 1):
        ws.row_dimensions[r].height = 15.75

    # ── Sloučení C6:C10 (prostor pro vlajku) ──────────────────
    ws.merge_cells('C6:C10')

    # ── D7:E7  Název pasu ─────────────────────────────────────
    ws.merge_cells('D7:E7')
    d7 = ws['D7']
    d7.value = 'Rostlinolékařský pas / Plant Passport'
    d7.font  = Font(bold=True, size=14)

    # ── D8:E8  Registrační číslo ──────────────────────────────
    ws.merge_cells('D8:E8')
    d8 = ws['D8']
    d8.value = 'B: CZ - 0550'
    d8.font  = Font(bold=True, size=14)

    # ── D9:E9 a D10:E10  (prázdné sloučené) ──────────────────
    ws.merge_cells('D9:E9')
    ws.merge_cells('D10:E10')

    # ── Řádek 12 — záhlaví sloupců ────────────────────────────
    center = Alignment(horizontal='center')
    for col, label in [('C', 'A:'), ('D', 'C:'), ('E', 'D:')]:
        cell           = ws[f'{col}12']
        cell.value     = label
        cell.font      = Font(bold=True, size=11)
        cell.alignment = center

    # ── Data od řádku 13 ──────────────────────────────────────
    font_data = Font(size=10)
    for i, plant in enumerate(plants):
        row = 13 + i
        ws.cell(row=row, column=3, value=plant.get('passport_name', '')).font = font_data
        ws.cell(row=row, column=4, value=plant.get('code', '')).font          = font_data
        ws.cell(row=row, column=5, value=plant.get('country', 'CZ')).font     = font_data

    # ── Ohraničení tabulky (řádky 12 až closing row) ──────────
    thin        = Side(style='thin')
    full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row + data rows + 1 prázdný závěrečný řádek
    closing_row = 13 + len(plants)   # empty row after last plant
    for row in range(12, closing_row + 1):
        for col in [3, 4, 5]:       # C, D, E
            ws.cell(row=row, column=col).border = full_border

    # ── Jeden souvislý rámeček celého pasu (C6:E{closing}) ────
    _outline(ws, 3, 6, 5, closing_row, thin)   # spojité levé/pravé i horní/dolní okraje
    for col in (3, 4, 5):                       # předěl hlavička | tabulka pod vlajkou
        _add_side(ws.cell(11, col), bottom=thin)
    for r in range(6, 12):                      # svislý předěl: vlajka | název pasu
        _add_side(ws.cell(r, 3), right=thin)
        _add_side(ws.cell(r, 4), left=thin)

    # ── EU vlajka v C6 ────────────────────────────────────────
    try:
        img        = XLImage(flag_path)
        img.width  = 181.9   # px — přesně dle vzoru (ručně upravený list 637)
        img.height = 105.6   # px
        img.anchor = 'C6'
        ws.add_image(img)
    except Exception as e:
        print(f"Varování: Nelze vložit vlajku: {e}")
