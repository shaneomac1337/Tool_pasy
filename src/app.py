import os
import tempfile
from datetime import date, datetime
import threading
import uuid
from pathlib import Path
import openpyxl
from flask import Flask, render_template, request, jsonify, send_file
from pdf_parser import PDFParser
from plant_matcher import PlantMatcher
from passport_generator import generate_excel
from pdf_generator import build_outputs, zip_outputs
from drive_uploader import get_status, upload_outputs, collect_output_files
from session import session
from paths import DATA_DIR, OUTPUT_DIR

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024

parser = PDFParser()

SARZE_PATH = DATA_DIR / 'sarze.xlsx'
matcher = PlantMatcher(str(SARZE_PATH))
print(f"✓ Šarže načtena: {len(matcher.entries)} rostlin")

UPLOAD_TMP = Path(tempfile.gettempdir()) / "rl_pasy_uploads"
UPLOAD_TMP.mkdir(exist_ok=True)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/parse', methods=['POST'])
def parse_pdfs():
    if 'files' not in request.files:
        return jsonify({'error': 'Žádné soubory nebyly nahrány'}), 400

    files = request.files.getlist('files')
    pdf_files = [f for f in files if f.filename.lower().endswith('.pdf')]
    if not pdf_files:
        return jsonify({'error': 'Žádné PDF soubory nenalezeny'}), 400

    saved_paths = []
    file_paths = {}
    for f in pdf_files:
        tmp_path = UPLOAD_TMP / f.filename
        f.save(str(tmp_path))
        saved_paths.append(str(tmp_path))
        file_paths[f.filename] = str(tmp_path)

    invoices = parser.parse_files(saved_paths)
    session.save(invoices, file_paths)

    return jsonify({
        'invoices':         [inv.to_dict() for inv in invoices],
        'total_invoices':   len(invoices),
        'total_plants':     sum(len(inv.plants) for inv in invoices),
        'files_processed':  len(saved_paths)
    })


@app.route('/api/match', methods=['POST'])
def match_plants():
    if not session.invoices():
        return jsonify({'error': 'Nejdřív nahraj faktury'}), 400

    result = []
    stats = {'exact': 0, 'fuzzy': 0, 'none': 0}

    for inv in session.invoices():
        plants_raw = [p.to_dict() for p in inv.plants]
        matched    = matcher.match_invoice_plants(plants_raw)
        for p in matched:
            stats[p['match_type']] += 1
        result.append({
            'number':      inv.number,
            'date':        inv.date,
            'customer':    inv.customer,
            'source_file': inv.source_file,
            'plants':      matched,
            'excluded':    inv.excluded,
        })

    return jsonify({
        'invoices':    result,
        'stats':       stats,
        'sarze_names': matcher.get_all_sarze_names(),
    })


@app.route('/api/search-plant', methods=['POST'])
def search_plant():
    """Ruční vyhledávání: uživatel zadá název, vrátí shodu ze šarže."""
    data  = request.get_json()
    query = (data or {}).get('query', '').strip()
    if not query:
        return jsonify({'error': 'Prázdný dotaz'}), 400

    result = matcher.search_by_name(query)
    return jsonify(result)


@app.route('/api/add-sarze', methods=['POST'])
def add_sarze():
    """Přidá novou položku do šarže (data/sarze.xlsx) a obnoví matcher."""
    global matcher
    data    = request.get_json() or {}
    name    = (data.get('name') or '').strip()
    code    = (data.get('code') or '').strip()
    country = (data.get('country') or '').strip() or 'CZ'
    if not name or not code:
        return jsonify({'error': 'Název i kód jsou povinné'}), 400

    # ponytail: bez zámku souboru — lokální nástroj pro jednoho uživatele;
    # při víceuživatelském provozu doplnit per-file zámek.
    wb = openpyxl.load_workbook(SARZE_PATH)
    ws = wb.active
    # Zapisuj hned za poslední NEPRÁZDNÝ řádek — exporty mívají na konci
    # stovky prázdných naformátovaných řádků a append() by zapsal až pod ně,
    # kde záznam v Excelu nikdo nenajde.
    last = ws.max_row
    while last > 1 and not any(cell.value for cell in ws[last]):
        last -= 1
    for col, value in enumerate([name, code, country], start=1):
        ws.cell(row=last + 1, column=col, value=value)
    wb.save(SARZE_PATH)

    matcher = PlantMatcher(str(SARZE_PATH))
    return jsonify(matcher.search_by_name(name))


def _validated_invoices():
    """Vrátí (invoices, None), nebo (None, chybová odpověď) při špatném těle."""
    data = request.get_json()
    if not data or 'invoices' not in data:
        return None, (jsonify({'error': 'Chybí data faktur'}), 400)
    invoices = data['invoices']
    if not invoices:
        return None, (jsonify({'error': 'Žádné faktury k exportu'}), 400)
    return invoices, None


@app.route('/api/generate-excel', methods=['POST'])
def generate_excel_route():
    """
    Přijme finální data z frontendu a vygeneruje Excel s pasy.
    Body: { "invoices": [ { "number", "customer", "date", "plants": [...] } ] }
    """
    invoices, error = _validated_invoices()
    if error:
        return error

    output_dir = OUTPUT_DIR
    try:
        excel_path = generate_excel(invoices, output_dir)
    except Exception as e:
        return jsonify({'error': f'Chyba při generování: {e}'}), 500

    return send_file(
        str(excel_path),
        as_attachment=True,
        download_name=excel_path.name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/generate-pdf', methods=['POST'])
def generate_pdf_route():
    """
    Přijme finální data a vygeneruje ZIP: pas + faktura per PDF,
    recipients.xlsx. Kopie zůstává ve výstupy/pasy_{datum}/.
    Body: { "invoices": [ { "number", "customer", "date", "plants": [...] } ] }
    """
    invoices, error = _validated_invoices()
    if error:
        return error

    out_dir = OUTPUT_DIR / f"pasy_{date.today().isoformat()}"
    try:
        result = build_outputs(
            invoices,
            session.index_by_number(),
            session.file_paths(),
            out_dir
        )
    except Exception as e:
        return jsonify({'error': f'Chyba při generování: {e}'}), 500

    buf = zip_outputs(result['files'])

    return send_file(
        buf,
        as_attachment=True,
        download_name=f"pasy_{date.today().isoformat()}.zip",
        mimetype='application/zip'
    )


_UPLOAD_LOCK = threading.Lock()
_UPLOAD_JOBS = {}
_ACTIVE_UPLOAD_JOB_ID = None


def _copy_upload_state(job_id: str) -> dict | None:
    with _UPLOAD_LOCK:
        state = _UPLOAD_JOBS.get(job_id)
        return dict(state) if state is not None else None


def _set_upload_state(job_id: str, **changes) -> None:
    with _UPLOAD_LOCK:
        if job_id in _UPLOAD_JOBS:
            _UPLOAD_JOBS[job_id].update(changes)


def _run_upload_job(job_id: str, date_str: str) -> None:
    global _ACTIVE_UPLOAD_JOB_ID

    def report(event):
        done = event.get('done', 0)
        total = event.get('total', 0)
        current = event.get('current', '')
        percent = int((done / total) * 100) if total > 0 else 0
        _set_upload_state(
            job_id, done=done, total=total, percent=percent, current=current)

    try:
        result = upload_outputs(date_str, progress_callback=report, max_workers=4)
        state = _copy_upload_state(job_id) or {}
        total = state.get('total', len(result['files']))
        _set_upload_state(
            job_id,
            status='done',
            done=total,
            total=total,
            percent=100,
            current='',
            folder_link=result['folder_link'],
            files=result['files'],
            error=None,
            finished_at=datetime.now().isoformat(timespec='seconds'),
        )
    except Exception as e:
        _set_upload_state(
            job_id,
            status='error',
            error=str(e),
            finished_at=datetime.now().isoformat(timespec='seconds'),
        )
    finally:
        with _UPLOAD_LOCK:
            if _ACTIVE_UPLOAD_JOB_ID == job_id:
                _ACTIVE_UPLOAD_JOB_ID = None


@app.route('/api/drive-status', methods=['GET'])
def drive_status_route():
    """Stav napojení na Google Drive (credentials + token)."""
    return jsonify(get_status())


@app.route('/api/upload-drive', methods=['POST'])
def upload_drive_route():
    """
    Spustí nahrávání výstupů dne na Google Drive: Excel do Rostlinné
    pasy/{rok}/, PDF do Rostlinné pasy/{rok}/pdfka/{datum}/.
    Body (volitelné): { "date": "YYYY-MM-DD" } — výchozí dnešek.
    """
    global _ACTIVE_UPLOAD_JOB_ID
    date_str = (request.get_json(silent=True) or {}).get('date') or date.today().isoformat()
    local_files = collect_output_files(date_str)
    if not local_files:
        return jsonify({'error': 'Složka výstupů neexistuje — nejprve vygenerujte PDF'}), 400

    with _UPLOAD_LOCK:
        if (
            _ACTIVE_UPLOAD_JOB_ID
            and _UPLOAD_JOBS.get(_ACTIVE_UPLOAD_JOB_ID, {}).get('status') == 'running'
        ):
            return jsonify({
                'error': 'Nahrávání už probíhá',
                'job_id': _ACTIVE_UPLOAD_JOB_ID,
                'status': 'running',
            }), 409

        job_id = uuid.uuid4().hex
        _UPLOAD_JOBS.clear()
        _UPLOAD_JOBS[job_id] = {
            'job_id': job_id,
            'status': 'running',
            'date': date_str,
            'started_at': datetime.now().isoformat(timespec='seconds'),
            'finished_at': None,
            'total': len(local_files),
            'done': 0,
            'percent': 0,
            'current': 'Čekám na Google Drive…',
            'folder_link': '',
            'files': [],
            'error': None,
        }
        _ACTIVE_UPLOAD_JOB_ID = job_id

    threading.Thread(
        target=_run_upload_job, args=(job_id, date_str), daemon=True).start()
    return jsonify({'job_id': job_id, 'status': 'running'}), 202


@app.route('/api/upload-drive/progress/<job_id>', methods=['GET'])
def upload_drive_progress_route(job_id):
    state = _copy_upload_state(job_id)
    if state is None:
        return jsonify({'error': 'Upload job nenalezen'}), 404
    return jsonify(state)


@app.route('/api/debug-tables', methods=['POST'])
def debug_tables():
    """Debug endpoint — vrátí surová data tabulek z prvního PDF."""
    import pdfplumber
    if 'file' not in request.files:
        return jsonify({'error': 'Žádný soubor'}), 400
    f = request.files['file']
    tmp_path = UPLOAD_TMP / f.filename
    f.save(str(tmp_path))

    pages_data = []
    with pdfplumber.open(str(tmp_path)) as pdf:
        for page_num, page in enumerate(pdf.pages[:3]):  # max 3 stránky
            tables = page.extract_tables()
            tables_data = []
            for t_idx, table in enumerate(tables):
                rows_data = []
                for r_idx, row in enumerate(table[:15]):  # max 15 řádků
                    rows_data.append({
                        'row_index': r_idx,
                        'cells': [str(c) if c is not None else 'None' for c in row]
                    })
                tables_data.append({
                    'table_index': t_idx,
                    'num_rows': len(table),
                    'num_cols': len(table[0]) if table else 0,
                    'rows': rows_data
                })
            pages_data.append({
                'page': page_num + 1,
                'num_tables': len(tables),
                'tables': tables_data
            })

    return jsonify({'pages': pages_data})

