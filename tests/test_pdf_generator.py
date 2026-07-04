# -*- coding: utf-8 -*-
"""Testy pro pdf_parser (page-run grouping) a pdf_generator (pas + merge)."""
import io

import openpyxl
import pdfplumber
import pytest
from pypdf import PdfReader
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

import pdf_generator  # noqa: F401  — side effect: registruje font 'Arial'
from pdf_generator import build_outputs, render_passport_pdf
from pdf_parser import Invoice, PDFParser

PLANTS = [{'passport_name': 'Dřín velkoplodý', 'code': 'CZ-123', 'country': 'CZ'}]


def _make_pdf(path, pages):
    """Vytvoří PDF; pages = list stránek, každá = list řádků textu.

    Font 'Arial' (registrovaný importem pdf_generator) je nutný, aby
    pdfplumber správně extrahoval české znaky ('č' ve 'Faktura č.').
    """
    c = canvas.Canvas(str(path), pagesize=A4)
    for lines in pages:
        c.setFont('Arial', 12)  # showPage resetuje stav — nastavit na každé stránce
        y = 750
        for line in lines:
            c.drawString(72, y, line)
            y -= 20
        c.showPage()
    c.save()


@pytest.fixture
def source_pdf(tmp_path):
    """Dvoustránková zdrojová faktura č. 111 (strana 2 = pokračování)."""
    path = tmp_path / 'faktura.pdf'
    _make_pdf(path, [
        ['Faktura č.: 111', 'E-mail: a@b.cz'],
        ['Pokračování položek bez hlavičky'],
    ])
    return path


def _parsed_index(source_pdf):
    return {'111': Invoice(number='111', email='a@b.cz',
                           source_file=source_pdf.name,
                           source_pages=[1, 2])}


# ---------------------------------------------------------------- parser

def test_parser_page_run_grouping(tmp_path):
    pdf_path = tmp_path / 'faktury.pdf'
    _make_pdf(pdf_path, [
        ['Faktura č.: 111', 'E-mail: a@b.cz'],
        ['Pokračování položek bez hlavičky'],
        ['Faktura č.: 112'],
    ])

    invoices = PDFParser()._parse_pdf(pdf_path)

    assert [inv.number for inv in invoices] == ['111', '112']
    inv111, inv112 = invoices
    assert inv111.source_pages == [1, 2]
    assert inv111.email == 'a@b.cz'
    assert inv112.source_pages == [3]
    assert inv112.email == ''


def test_parse_items_excludes_gift_voucher():
    """Řádky 'Dárkový poukaz' a 'Přidaný produkt' se při parsování vynechají."""
    table = [
        ['() Dárkový poukaz v hodnotě 1000 Kč', '1', '1 000', '1 000'],
        ['() Přidaný produkt', '1', '50', '50'],
        ['() Agave parryi var. parryi', '3', '120', '360'],
    ]
    parser = PDFParser()
    plants = parser._parse_items(table)

    assert [p.name for p in plants] == ['Agave parryi var. parryi']
    assert parser._is_non_plant('Dárkový poukaz v hodnotě 1000 Kč') is True
    assert parser._is_non_plant('Agave parryi var. parryi') is False


# ---------------------------------------------------------------- build_outputs

def test_merge_with_source_invoice(tmp_path, source_pdf):
    out_dir = tmp_path / 'out'
    result = build_outputs(
        [{'number': '111', 'customer': 'X', 'date': '', 'plants': PLANTS}],
        _parsed_index(source_pdf),
        {source_pdf.name: str(source_pdf)},
        out_dir,
    )

    merged = out_dir / '111.pdf'
    assert merged in result['files']
    # 1 stránka pasu + 2 stránky zdrojové faktury
    assert len(PdfReader(merged).pages) == 1 + 2
    assert result['warnings'] == []
    assert not (out_dir / 'varovani.txt').exists()


def test_fallback_missing_source(tmp_path):
    out_dir = tmp_path / 'out'
    result = build_outputs(
        [{'number': '999', 'customer': 'Y', 'date': '', 'plants': PLANTS}],
        {},   # číslo není v parsed_index
        {},
        out_dir,
    )

    passport_only = out_dir / '999.pdf'
    assert passport_only.exists()
    assert len(PdfReader(passport_only).pages) == 1
    assert len(result['warnings']) == 1
    assert '999' in result['warnings'][0]
    warn_path = out_dir / 'varovani.txt'
    assert warn_path.exists()
    assert warn_path.read_text(encoding='utf-8') == result['warnings'][0]


def test_empty_plants_skipped(tmp_path, source_pdf):
    out_dir = tmp_path / 'out'
    result = build_outputs(
        [{'number': '111', 'customer': 'X', 'date': '', 'plants': []}],
        _parsed_index(source_pdf),
        {source_pdf.name: str(source_pdf)},
        out_dir,
    )

    assert not (out_dir / '111.pdf').exists()
    assert result['warnings'] == []
    ws = openpyxl.load_workbook(out_dir / 'recipients.xlsx').active
    assert ws.max_row == 1  # jen hlavička, žádný příjemce


def test_recipients_xlsx(tmp_path, source_pdf):
    out_dir = tmp_path / 'out'
    build_outputs(
        [{'number': '111', 'customer': 'X', 'date': '', 'plants': PLANTS}],
        _parsed_index(source_pdf),
        {source_pdf.name: str(source_pdf)},
        out_dir,
    )

    ws = openpyxl.load_workbook(out_dir / 'recipients.xlsx').active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ('Email', 'Attachment')
    assert rows[1] == ('a@b.cz', '111.pdf')
    assert len(rows) == 2


# ---------------------------------------------------------------- rendering

def test_render_passport_pdf(tmp_path):
    buf = io.BytesIO()
    render_passport_pdf(PLANTS, buf)

    buf.seek(0)
    assert len(PdfReader(buf).pages) == 1

    buf.seek(0)
    with pdfplumber.open(buf) as pdf:
        text = pdf.pages[0].extract_text() or ''
    assert 'Dřín velkoplodý' in text
    assert 'Rostlinolékařský pas / Plant Passport' in text
    assert 'B: CZ - 0550' in text
